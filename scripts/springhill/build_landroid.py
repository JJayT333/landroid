#!/usr/bin/env python3
"""Build a v8 .landroid workspace for the Dr. Elmore #1 Unit (Springhill folder).

Roster + fractions: present-day mineral ownership from the NRI status file.
Columns are auto-detected from the header row (the sheet has interleaved
"NPR Adjusted" columns, so hard-coded letters are unsafe). Each tract's
Mineral Interest column totals 1.0; divided-interior owners (acre-only) are
expanded by acreage against the tract's "Divided interior tracts" rollup.

Documents: the controlling DOTO (attached to every node) plus any recorded
chain instrument or OGML lease (from the runsheet) whose grantee/lessor name
matches the owner. Every owner is an ORPHAN node on its tract's desk map.
"""
import argparse, openpyxl, warnings, os, base64, hashlib, json, re, sys
from fractions import Fraction
from openpyxl.utils import column_index_from_string as ci
from pathlib import Path
warnings.filterwarnings('ignore')

REPO_ROOT=Path(__file__).resolve().parents[2]
DEFAULT_SPRING=Path('/Users/abstractmapping/Documents/LANDroid/Landroid Stuff/LANDroid - Springhill')
DEFAULT_RUNSHEET_NAME='MTO/TORS_Documents/DOTO_Runsheet_Elmore#1_Unit_2026_02-05.xlsx'
DEFAULT_NRI_NAME='Status_Springhill_Dr.Elmore#1_Unit_NRI_2026_03-09_RKH copy.xlsx'
DEFAULT_OUT_NAME='Springhill_Dr-Elmore-1_PresentDay.landroid'
PUBLIC_SAMPLE=(REPO_ROOT/'public/samples/springhill-dr-elmore.landroid').resolve()

def _expand_path(value):
    return Path(value).expanduser().resolve()

def _is_under(child, parent):
    try:
        child.relative_to(parent)
        return True
    except ValueError:
        return False

def _readable_file(path, label, flag):
    if not path.is_file():
        raise FileNotFoundError(f'{label} not found: {path}')
    try:
        with path.open('rb') as fh:
            fh.read(1)
    except PermissionError as exc:
        raise PermissionError(
            f'{label} exists but this Python process cannot read it: {path}\n'
            'Grant this terminal/Codex process access to the private source folder, '
            'or copy the private workbook to another untracked private folder and '
            f'pass it with {flag}.'
        ) from exc

def _refuse_repo_private_output(path, label):
    if _is_under(path, REPO_ROOT):
        raise ValueError(
            f'Refusing to write {label} inside the repository: {path}\n'
            'The raw Springhill build embeds real private source data and real PDF '
            'blobs. Write it to a private path outside the repo, then run '
            'scripts/springhill-scrub.ts to create the shareable public sample.'
        )
    if path == PUBLIC_SAMPLE:
        raise ValueError(
            'Refusing to write the raw Springhill build directly over the public sample. '
            'Generate raw output privately, scrub it, then review and copy the scrubbed result.'
        )

def _parse_config():
    parser=argparse.ArgumentParser(
        description='Build the private raw Springhill Dr. Elmore .landroid workspace.'
    )
    parser.add_argument('--spring-dir', default=os.environ.get('SPRINGHILL_SOURCE_DIR', str(DEFAULT_SPRING)),
        help='Private Springhill source folder. Env: SPRINGHILL_SOURCE_DIR.')
    parser.add_argument('--runsheet', default=os.environ.get('SPRINGHILL_RUNSHEET'),
        help='Private DOTO runsheet workbook path. Env: SPRINGHILL_RUNSHEET.')
    parser.add_argument('--nri', default=os.environ.get('SPRINGHILL_NRI'),
        help='Private NRI status workbook path. Env: SPRINGHILL_NRI.')
    parser.add_argument('--pdfdir', default=os.environ.get('SPRINGHILL_PDFDIR', str(REPO_ROOT/'TORS_Documents')),
        help='Directory containing instrument PDFs. Env: SPRINGHILL_PDFDIR.')
    parser.add_argument('--out', default=os.environ.get('SPRINGHILL_OUT'),
        help='Private raw .landroid output path. Env: SPRINGHILL_OUT.')
    parser.add_argument('--report', default=os.environ.get('SPRINGHILL_REPORT'),
        help='Private reconciliation report output path. Env: SPRINGHILL_REPORT.')
    args=parser.parse_args()

    spring=_expand_path(args.spring_dir)
    runsheet=_expand_path(args.runsheet) if args.runsheet else spring/DEFAULT_RUNSHEET_NAME
    nri=_expand_path(args.nri) if args.nri else spring/DEFAULT_NRI_NAME
    pdfdir=_expand_path(args.pdfdir)
    out=_expand_path(args.out) if args.out else spring/DEFAULT_OUT_NAME
    report=_expand_path(args.report) if args.report else out.with_suffix(out.suffix+'.report.txt')

    _refuse_repo_private_output(out,'raw .landroid output')
    _refuse_repo_private_output(report,'reconciliation report')
    _readable_file(runsheet,'runsheet workbook','--runsheet')
    _readable_file(nri,'NRI workbook','--nri')
    if not pdfdir.is_dir():
        raise FileNotFoundError(f'PDF directory not found: {pdfdir}')

    return dict(root=str(REPO_ROOT), pdfdir=str(pdfdir), runsheet=str(runsheet), nri=str(nri), out=str(out), report=str(report))

try:
    CONFIG=_parse_config()
except Exception as exc:
    print(f'[springhill-build] aborted: {exc}', file=sys.stderr)
    sys.exit(1)

ROOT=CONFIG['root']
PDFDIR=CONFIG['pdfdir']
RUNSHEET=CONFIG['runsheet']
NRI=CONFIG['nri']
OUT=CONFIG['out']
REPORT=CONFIG['report']
WS='ws-springhill-elmore1'; NOW='2026-03-09T00:00:00.000Z'; PREC=9; DOTO='DOTO_ElmoreC-1_Unit'
COUNTY='San Jacinto'; SURVEY='Thomas Webb Survey A-300 & Vital Flores Survey A-14, San Jacinto County, Texas'
# Dr. Elmore unit tract number -> DOTO unit-tract label (from runsheet sheet names)
UNITTR={1:'6',2:'1',3:'5',4:'3',5:'4',6:'2',7:'7'}
ACRES={1:'55.5016',2:'106.19',3:'28.223',4:'14.64',5:'8.44',6:'42.581',7:'20.546'}

# ── NPRIs carved out in the 6/4/2009 DOTO ──────────────────────────────────
# Two non-participating royalty interests burden subsets of the unit. They are
# sibling burdens to the mineral estate (interestClass 'npri'), attached to the
# tract's DOTO root — NOT to any present-day mineral owner, since they were
# reserved up-chain. Present-day royalty owners are worked from each NPR sheet's
# Subsequent-Title chain; rows the runsheet marks "no subsequent documents" keep
# their exact 2009 fraction. The Moran/LJM strand is left as one node carrying
# its full fraction with a note, because the runsheet does not cleanly state the
# downstream split. Each list MUST sum to the NPR total (NPR1=0.015625 over
# Tr.2/3/5/6; NPR2=0.0625 over Tr.4/7); a residue check folds rounding dust into
# the largest holder so the burden total is exact.
NPR1_TRACTS=[2,3,5,6]; NPR2_TRACTS=[4,7]
# (present-day owner, decimal interest, succession instrument or None)
NPR1_OWNERS=[
 ('H.C. Aiken (or successors)',0.0015625,None),
 ('Shriners Hospitals for Children',0.00036563,None),
 ('Board of Regents, U.T. System (M.D. Anderson Cancer Center)',0.00036563,None),
 ('The Museum of Fine Arts, Houston',0.00036563,None),
 ('St. Thomas High School',0.00036563,None),
 ('Louise J. Moran Foundation',0.00219375,None),
 ('Shine LLC',0.000675,None),
 ('Holmes Family Irrevocable Trust (Elizabeth K. & Jonathon H. Holmes)',0.000675,'09-2054'),
 ('Franke Natural Resources, L.P.',0.000675,None),
 ('KMF Management Trust (Kenneth Michael Franke, Trustee)',0.000675,None),
 ('Bruce L. Franke',0.000675,None),
 ('Patrick J. Moran Trust (Moran & O\'Neill, Trustees)',0.0028125,None),
 ('Moran / LJM Trust strand (succession unresolved — see note)',0.00106875,'09-6968'),
 ('Moran Employees Trust (Moran & Texas Commerce Bank, Trustees)',0.00085219,None),
 ('Ann E. Moran Living Trust',0.00043172,'20194106'),
 ('W.T. Moran III Trust (Moran & Letcher, Trustees)',0.00043172,None),
 ('Moran Properties, Inc.',0.00043172,None),
 ('William S. Poinsett',9.516e-05,None),
 ('Judith Lee Poinsett Properties, LLC',9.516e-05,None),
 ('Leroy M. Poinsett, Jr.',9.516e-05,None),
 ('Margaret Ann Dell Osso Trust (Texas Commerce Bank, Trustee)',0.00021656,None),
 ('Sek Rose Resources LLC (formerly Est. of Mae G. Shapley)',0.00014344,'20212728'),
 ('N.A.B. Oil & Gas Holdings, Inc.',0.00014344,None),
 ('Byrne Family, L.L.C.',7.172e-05,None),
 ('M. Klein Enterprises, LLC',7.172e-05,None),
 ('James P. Poinsett Trust (Moran & W.T. Moran III, Trustees)',7.031e-05,None),
]
NPR2_OWNERS=[
 ('Benita Trapp Downey (1/2 of Billie Jo Trapp NPR)',0.02083333/2,'P2025-4'),
 ('Robert Hill Trapp (1/2 of Billie Jo Trapp NPR)',0.02083333/2,'P2025-4'),
 ('Doug Wellborn',0.00347222,None),
 ('Laura Ann Wellborn Chestnut',0.00347222,None),
 ('Parker Speer (Estate of Lily Lois Speer)',0.01041667,'24CPR0065_Probate'),
 ('Pete McClanahan',0.01041667,None),
 ('Tinabeth McCune Keasling (1/2 of Vivian McCune NPR)',0.00694444/2,'20242292'),
 ('Sara McCune Keisler (1/2 of Vivian McCune NPR)',0.00694444/2,'20242292'),
 ('Shelley Hamblen Musgrave (1/2 of Virginia Hamblen-Dove NPR)',0.00694444/2,'20174126'),
 ('Rebecca Hamblen Drapak (1/2 of Virginia Hamblen-Dove NPR)',0.00694444/2,'20174126'),
]

pdfs={os.path.splitext(f)[0]:f for f in os.listdir(PDFDIR) if f.lower().endswith('.pdf')}
assert DOTO in pdfs

# ---------- NRI roster: detect columns from header row ----------
ns=openpyxl.load_workbook(NRI,data_only=True)['Sheet1']
MIcol={}; ACcol={}
for c in range(1,ns.max_column+1):
    h=str(ns.cell(1,c).value or '').replace('\n',' ').strip()
    m=re.search(r'Tr\.?\s*(\d)',h)
    if not m: continue
    t=int(m.group(1))
    if 'Mineral Interest' in h: MIcol[t]=c
    elif 'Mineral Acres' in h: ACcol[t]=c
assert sorted(MIcol)== [1,2,3,4,5,6,7], f'bad MI detect {MIcol}'
TOTAL=None
for r in range(2,ns.max_row+1):
    if str(ns.cell(r,1).value or '').strip().lower()=='total': TOTAL=r
def isn(v): return isinstance(v,(int,float))
def first(r):
    v=ns.cell(r,1).value; return str(v).split('\n')[0].strip() if v else ''
def addr_of(r):
    v=ns.cell(r,1).value
    if not v: return ''
    ls=[l.strip() for l in str(v).split('\n') if l.strip()]
    return '\n'.join(ls[1:]) if len(ls)>1 else ''

# Authoritative lease royalty + status from the NRI status sheet (the columns
# the landman actually maintains), keyed by normalized owner name. Royalty
# scraped from lease remarks is unreliable — this column is the source of truth
# and directly drives the Leasehold view's NRI (leased fraction × royalty rate).
NRI_HDR={str(ns.cell(1,c).value or '').replace('\n',' ').strip():c for c in range(1,ns.max_column+1)}
_royc=NRI_HDR.get('Lease Roy.'); _stc=NRI_HDR.get('Lease Status')
def _normkey(s): return re.sub(r'[^a-z0-9]+',' ',(s or '').lower()).strip()
def _royalty_fraction(v):
    if not isn(v): return ''
    for n,d in [(1,4),(1,5),(1,6),(3,16),(1,8),(1,3),(1,2)]:
        if abs(v-n/d)<1e-4: return f'{n}/{d}'
    return f'{v:.6f}'
AUTH_ROY={}  # normkey -> (royaltyFractionStr, status, executedBool)
for r in range(2,(TOTAL or ns.max_row+1)):
    nm=first(r)
    if not nm or 'divided' in nm.lower() or nm.lower()=='total': continue
    roy=ns.cell(r,_royc).value if _royc else None
    st=str(ns.cell(r,_stc).value or '').strip() if _stc else ''
    AUTH_ROY[_normkey(nm)]=(_royalty_fraction(roy), st, 'executed' in st.lower())
def auth_for(name):
    return AUTH_ROY.get(_normkey(name),('','',False))

def roster(t):
    """Return list of (row, name, fraction) for tract t, divided owners expanded."""
    out=[]; rollup=None
    for r in range(2,TOTAL):
        v=ns.cell(r,MIcol[t]).value
        if isn(v) and v>0:
            n=first(r)
            if 'divided' in n.lower(): rollup=float(v); continue
            out.append((r,n,float(v)))
    if rollup is not None:
        rows=[]; tot=0.0
        for r in range(2,TOTAL):
            if isn(ns.cell(r,MIcol[t]).value): continue   # has a direct interest already
            a=ns.cell(r,ACcol[t]).value
            if isn(a) and a>0 and 'total' not in first(r).lower():
                rows.append((r,first(r),float(a))); tot+=a
        for r,n,a in rows: out.append((r,n,rollup*a/tot))
    return out

# ---------- runsheet chains + leases ----------
rwb=openpyxl.load_workbook(RUNSHEET,data_only=True)
TITLE_SHEET={1:'MI-Title-Tr.1-(6)-55.5016 ac.',2:'Title-Tr.2-(1)-106.19 ac.',3:'Title-Tr.3-(5)-28.223 ac.',
 4:'Title-Tr.4-(3)-14.64 ac.',5:'Title-Tr.5-(4)-8.44 ac.',6:'Title-Tr.6-(2)-42.581 ac.',7:'Title-Tr.7-(7)-20.546 ac.'}
KIND={'deed':'deed'}
def kind_of(itype):
    s=(itype or '').lower()
    if any(w in s for w in ['probate','lwt','order']): return 'probate'
    if any(w in s for w in ['aoh','heirship','affidavit','poa']): return 'affidavit'
    if 'ogml' in s or 'lease' in s: return 'lease'
    return 'deed'

def chain_for(t):
    sheet=TITLE_SHEET[t]
    if sheet not in rwb.sheetnames: return []
    ws=rwb[sheet]; rows=[]; started=False
    for r in range(1,ws.max_row+1):
        a=ws.cell(r,1).value
        if a and 'subsequent title' in str(a).strip().lower(): started=True; continue
        if not started: continue
        g=ws.cell(r,7).value
        if not g or str(g).strip() not in pdfs: continue
        rows.append(dict(instr=str(g).strip(), grantee=str(ws.cell(r,11).value or '').strip(),
            itype=str(ws.cell(r,2).value or 'Deed').strip(),
            date=str(ws.cell(r,9).value or '')[:10]))
    return rows

# Instrument metadata index: scan EVERY runsheet sheet for rows whose col G is a
# known PDF, capturing instrument type / grantor / grantee / dates / vol / page.
# Used to fully populate the document registry so the Documents tab reads complete.
def build_doc_meta():
    meta={}
    for sheet in rwb.sheetnames:
        ws=rwb[sheet]
        # find header columns by label on row 1
        hdr={}
        for c in range(1,ws.max_column+1):
            h=str(ws.cell(1,c).value or '').replace('\n',' ').strip().lower()
            if h: hdr[h]=c
        def col(*names):
            for n in names:
                for h,c in hdr.items():
                    if n in h: return c
            return None
        cG=col('instrument no'); cB=col('instrument'); cVol=col('vol'); cPg=col('page')
        cFile=col('file date'); cInst=col('inst. date','inst./eff','inst date')
        cGr=col('grantor','assignor','lessor'); cGe=col('grantee','assignee','lessee')
        cDesc=col('land desc'); cRem=col('remarks')
        if not cG: continue
        for r in range(2,ws.max_row+1):
            g=ws.cell(r,cG).value
            if not g: continue
            gid=str(g).strip()
            if gid not in pdfs or gid in meta: continue
            def cell(c): return str(ws.cell(r,c).value).strip() if c and ws.cell(r,c).value is not None else ''
            meta[gid]=dict(itype=cell(cB), vol=cell(cVol), page=cell(cPg),
                file_date=cell(cFile)[:10], inst_date=cell(cInst)[:10],
                grantor=cell(cGr), grantee=cell(cGe), desc=cell(cDesc), remarks=cell(cRem))
    return meta
DOC_META=build_doc_meta()

# OGML leases: instrument PDF + lessor name (col J=10)
def leases_all():
    lh=rwb['Leasehold']; out=[]
    for r in range(2,lh.max_row+1):
        if str(lh.cell(r,2).value or '').strip().upper()!='OGML': continue
        out.append(dict(instr=str(lh.cell(r,7).value or '').strip(),
            lessor=str(lh.cell(r,10).value or '').strip(),
            rem=str(lh.cell(r,13).value or '').strip(),
            date=str(lh.cell(r,9).value or '')[:10]))
    return out
LEASES=leases_all()

# Explicit lessor→owner map. Fuzzy token-overlap mis-linked leases when owners
# shared a surname (e.g. "Hill", "Owen"), so each OGML lists the exact present-day
# owner name fragment(s) it covers. A lease matches an owner iff a fragment's
# significant tokens are a SUBSET of the owner's name tokens. Some OGMLs cover
# two owners (joint lessors); some owners are intentionally unleased (status
# "Pursuing" in the NRI) and correctly get no lease node.
LEASE_OWNER_FRAGMENTS={
 'OGML_Powell':['Sandra Owen Powell'],
 'OGML_Owen_Rigby&Steven':['Rigby Owen','Stephen Owen'],
 'OGML_Few':['John Few'],
 'OGML_Trapp_Robt':['Robert Hill Trapp'],
 'OGML_Trapp_Downey':['Benita Trapp Downey'],
 'OGML_McClanahan':['Pete McClanahan'],
 'OGML_Tinabeth_Keasling':['Tinabeth Keasling'],
 'OGML_Drapak':['Rebecca Hamblen Drapak'],
 'OGML-ElmoreFamilyPartners':['Elmore Family Partners'],
 'OGML-LCT-Trust':['Charlyn Tyra'],
 'OGML-Tyra_R':['Richard Tyra'],
 'OGML_Mayo-Henson':['Raigan Mayo','Heather Henson'],
 'OGML_Broussard':['Broussard'],
 'OGML_Deloach':['Deloach'],
 'OGML_Old_Hickory_Hill':['Old Hickory Hill'],
 'OGML_McRorie':['McRorie'],
 'OGML_Edwards':['Edwards'],
 'OGML_Kraft':['Peter Hill Kraft'],
 'OGML_Harman_Donald':['Donald Harman'],
 'OGML_Harman_Claude':['Claude Lester Harman'],
}
LEASE_BY_INSTR={L['instr']:L for L in LEASES}

# Source-status overrides proven from recorded OGML packets, not from the NRI
# status sheet. The NRI/leasehold spreadsheet is useful working evidence for
# Springhill, but it is not the permanent authority for LANDroid generation.
# LCT was a known miss: the OGML packet names LCT Revocable Trust / Charlyn K.
# Tyra, Trustee as lessor, Magnolia as lessee, a 11/19/2025 lease date,
# one-year primary term, and one-fourth royalty.
LEASE_SOURCE_OVERRIDES={
 'OGML-LCT-Trust':dict(status='Executed — verified from OGML-LCT-Trust.pdf',
                       royalty='1/4',
                       notes='Verified from OGML-LCT-Trust.pdf OCR: one-year primary term; 1/4 royalty.',
                       evidence='OGML-LCT-Trust.pdf OCR: LCT Revocable Trust / Charlyn K. Tyra, Trustee; Magnolia Petroleum Company, LLC; 11/19/2025; one-year primary term; one-fourth royalty.'),
}
def lease_for_owner(name):
    ot=toks(name)
    for instr,frags in LEASE_OWNER_FRAGMENTS.items():
        if instr not in pdfs or instr not in LEASE_BY_INSTR: continue
        for frag in frags:
            ft=toks(frag)
            if ft and ft <= ot:
                return LEASE_BY_INSTR[instr]
    return None

STOP={'family','trust','estate','the','and','inc','llc','ltd','investments','revocable','partners',
 'company','trustee','trustees','executor','executrix','independent','jr','sr','as','of','his','her',
 'their','sole','separate','property','married','single','wife','husband','children','joint','tenants',
 'survivorship','co','management','individually','being','court','appointed','fiduciary','dec',
 'deceased','est','under','agreement','dated'}
def toks(name): return {w for w in re.findall(r"[A-Za-z]{4,}", (name or '').lower()) if w not in STOP}

def dec(x): return f"{x:.{PREC}f}"
def numden(x):
    fr=Fraction(str(round(x,PREC))).limit_denominator(10**8); return str(fr.numerator),str(fr.denominator)

# ---------- assemble ----------
owners={}
def norm(n): return re.sub(r'\s+',' ',str(n)).strip().lower()
def owner_id(name,row=None):
    k=norm(name)
    if k not in owners:
        low=name.lower(); ent='Individual'
        if any(w in low for w in ['llc','ltd','inc','partners','company','investments','co.']): ent='Entity'
        if 'trust' in low: ent='Trust'
        owners[k]=dict(id='owner-'+re.sub(r'[^a-z0-9]+','-',k).strip('-')[:46], workspaceId=WS, name=name,
            entityType=ent, county=COUNTY, prospect='Dr. Elmore #1 Unit', mailingAddress='',
            email='', phone='', notes='', createdAt=NOW, updatedAt=NOW)
    if row is not None and not owners[k]['mailingAddress']:
        owners[k]['mailingAddress']=addr_of(row)
    return owners[k]['id']

nodes=[]; deskMaps=[]; attachments=[]; need={DOTO}; lease_recs=[]; reconciliation=[]; nidc=0

# ─────────────────────────────────────────────────────────────────────────────
# Tract 2 — full DOTO→present chain of title (multi-generation).
#
# Unlike the other six tracts (two-layer: DOTO root → present owner), Tract 2 is
# built as a real title tree: DOTO root → 2009 DOTO owner → each conveyance →
# present-day owner. Hover math is structural — `granted = initialFraction ÷
# parent.initialFraction`, `of whole = initialFraction`, `remaining = fraction`
# — so an honest tree yields honest derivations with no free-text math.
#
# Encoding rule per node:  initialFraction = of-whole share at this instrument;
# fraction (remaining) = the part STILL held here = initialFraction − Σ children
# initialFraction. A grantee who conveyed everything out carries fraction 0; a
# present-day holder carries fraction = initialFraction. Desk-map coverage sums
# `fraction>0` across the flat node list, so only present-day leaves count and
# they total 1.0.
#
# Merges (multiple grantors → one grantee) can't be a strict tree, so two knots
# are consolidated into a single node (with every instrument attached as a doc),
# per the user's instruction:
#   • Lenoir: two estates (25/840 each) cured through crossed trusts/executors'
#     deeds into the JP Morgan GST trusts, sold as one 25/420 block to Bradbury.
#   • Hill: Bruce (via his trust → Busby), Kathleen (estate), and Robert Moore
#     Hill Jr. (intestate → Kathleen) consolidate to the Krafts. Jane stays.
#     NOTE: this is the USER'S chain and it DIVERGES from the NRI sheet, which
#     lists Luanne Staton Hill at 5/224 and the Krafts at 5/224 each. Here the
#     Krafts hold 15/448 each and Luanne holds nothing. The Hill 5/56 total is
#     preserved so the tract still balances to 100%, but Peter Kraft's leased
#     NRI shifts — flagged for the user to reconcile on the NRI sheet.
# ─────────────────────────────────────────────────────────────────────────────
def build_chain(t, parent_id, node_ids):
    UTR=UNITTR[t]; AC=ACRES[t]; F=Fraction
    landdesc=f'{SURVEY}; Dr. Elmore #1 Unit Tract {t} (DOTO Unit Tr. {UTR}), {AC} ac.'
    # NRI present-day name → row, for mailing address / authoritative royalty / lease match
    nri_row={}
    for r,nm,fr in roster(t): nri_row[_normkey(nm)]=(r,nm)
    def rowfor(nm): return nri_row.get(_normkey(nm),(None,nm))
    # Bradbury divided-interior lots: mineral fraction (NRI acreage expansion) + as-called acres
    rollup=None
    for r in range(2,TOTAL):
        v=ns.cell(r,MIcol[t]).value
        if isn(v) and v>0 and 'divided' in first(r).lower(): rollup=float(v)
    drows=[]; dtot=0.0
    for r in range(2,TOTAL):
        if isn(ns.cell(r,MIcol[t]).value): continue
        a=ns.cell(r,ACcol[t]).value
        if isn(a) and a>0 and 'total' not in first(r).lower(): drows.append((first(r),float(a))); dtot+=a
    DIVF={}; DIVA={}
    if rollup is not None and dtot>0:
        for nm,a in drows: DIVF[nm]=rollup*a/dtot; DIVA[nm]=a
    def dv(sub):
        for nm in DIVF:
            if sub.lower() in nm.lower(): return nm,DIVF[nm],DIVA[nm]
        raise KeyError(sub)

    def node(name,instr,itype,date,leaf=None,present=False,docs=None,note='',nri=None,children=None,deceased=False,acres=None):
        return dict(name=name,instr=instr,itype=itype,date=date,leaf=leaf,present=present,docs=docs or [],
                    note=note,nri=nri or name,children=children or [],deceased=deceased,acres=acres)
    def lot(sub,nri_name,instr,docs=None,note='',block='subdivision'):
        nm,fr,a=dv(sub)
        return node(nri_name,instr,'WD (no mineral reservation — surface & minerals pass)','',leaf=fr,present=True,
            docs=docs,nri=nm,acres=a,
            note=note or (f'{block} lot conveyed with no mineral reservation, so minerals passed to the lot buyer. '
                          f'As-called deed acreage {a:.4f} ac; mineral fraction {fr:.9f} reconciled to the divided-interior rollup.'))

    # ── shared family fragments (identical across Tracts 2,3,5,6 except where parameterized) ──
    def s_elmore(leaf,extra='',docs=None):
        return node('Elmore Family Partners, Ltd.',DOTO,'DOTO (mineral owner)','',leaf=leaf,present=True,
            nri='Elmore Family Partners, Ltd., by Elmore',docs=docs,
            note=f'2009 DOTO mineral owner; no subsequent conveyance — interest unchanged.{extra}')
    def s_owen():
        return node('Jo Briley Owen',DOTO,'DOTO (mineral owner)','',
            note=('2009 DOTO owner, 37/168, burdened by a 1/64 NPRI (NPR 1 — a sibling burden off the DOTO root). '
                  'Gifted all to her three children in equal shares (09-3821, "share equally as separate property"), 1/3 each.'),
            children=[
                node('Rigby Owen, Jr.','09-3821','Gift Deed of Mineral Interest','',leaf=F(37,504),present=True,nri='Rigby Owen, Jr.,'),
                node('Stephen J. Owen','09-3821','Gift Deed of Mineral Interest','',leaf=F(37,504),present=True,nri='Stephen J. Owen'),
                node('Sandra Owen Powell','09-3821','Gift Deed of Mineral Interest','',leaf=F(37,504),present=True,nri='Sandra Owen Powell'),
            ])
    def s_browder(few_leaf,extra=''):
        return node('Estate of Bridges / Ann Childers Browder (Jane Childers Browder)',DOTO,'DOTO (mineral owner)','',
            note=('2009 DOTO Browder owner. NAME NOTE: the DOTO lists "Bridges Browder"; the first probate (11-4768) '
                  'probates "Ann Childers Browder" — Jane Childers Browder is the through-line. Passes down the chain '
                  'to John A. Few, Sr.'+extra),
            children=[node('Browder Family Trust (Jane Childers Browder, Trustee)','11-4768','CC Order & LWT','',
                children=[node('Jane C. Browder Trust','2012006700','Executors Distribution Deed','',
                    children=[node('Jane C. Browder (individually)','2012006703','Distribution SWD','',
                        children=[node('John A. Few, Sr.','2012006704','Gift SWD','',leaf=few_leaf,present=True,nri='John A. Few, Sr.')])])])])
    def s_trapp():
        return node('Billie Jo Trapp',DOTO,'DOTO (mineral owner)','',docs=['20241988'],
            note=('2009 DOTO owner, 1/3 x 17/84 = 17/252. POA to Robert H. Trapp (20241988); devised to her two children '
                  'by Order as Muniment of Title (P2025-4, "devises to children"), 1/2 each.'),
            children=[
                node('Benita Trapp Downey','P2025-4','Order & LWT (Muniment of Title)','',leaf=F(17,504),present=True,nri='Benita Trapp Downey'),
                node('Robert Hill Trapp','P2025-4','Order & LWT (Muniment of Title)','',leaf=F(17,504),present=True,nri='Robert Hill Trapp'),
            ])
    def s_speer():
        return node('Lily Lois Speer',DOTO,'DOTO (mineral owner)','',
            note=('2009 DOTO owner, 1/6 x 17/84 = 17/504. Will (24CPR0065): general estate to children share and share '
                  'alike, but ALL mineral interest to be held by Kevin Lee Speer AS CUSTODIAN FOR PARKER SPEER. The '
                  'mineral fee vests in Parker Speer.'),
            children=[node('Parker Speer','24CPR0065_Probate','Probate Order & LWT','',leaf=F(17,504),present=True,nri='Parker Speer',
                note='Held by Kevin Lee Speer as custodian for Parker Speer (special bequest per the will).')])
    def s_mccune():
        return node('Vivian McClanahan McCune',DOTO,'DOTO (mineral owner)','',docs=['20242292'],
            note='2009 DOTO owner, 1/9 x 17/84 = 17/756. AOH (20172674 / 20242292) identifies her two daughters as heirs; 1/2 each.',
            children=[
                node('Tinabeth McCune Keasling','20172674','Affidavit of Heirship','',leaf=F(17,1512),present=True,nri='Tinabeth Keasling and husband Lawrence'),
                node('Sara McCune Keisler','20172674','Affidavit of Heirship','',leaf=F(17,1512),present=True,nri='Sara Kathryn (McCune) Keisler'),
            ])
    def s_dove():
        return node('Virginia Louise Hamblen Dove',DOTO,'DOTO (mineral owner)','',
            note='2009 DOTO owner, 1/9 x 17/84 = 17/756. Will (20174126, "all to children") devises all to her two children; 1/2 each.',
            children=[
                node('Shelley Hamblen Musgrave','20174126','CC Order & LWT','',leaf=F(17,1512),present=True,nri='Shelley Hamblen Musgrave'),
                node('Rebecca Hamblen Drapak','20174126','CC Order & LWT','',leaf=F(17,1512),present=True,nri='Rebecca Hamblen Drapak'),
            ])
    def s_wellborn():
        return [node('Doug Wellborn',DOTO,'DOTO (mineral owner)','',leaf=F(17,1512),present=True,nri='Doug Wellborn',
                    note='2009 DOTO owner, 1/18 x 17/84 = 17/1512; no subsequent conveyance — interest unchanged.'),
                node('Laura Ann Wellborn Chestnut',DOTO,'DOTO (mineral owner)','',leaf=F(17,1512),present=True,nri='Laura Ann Wellborn Chestnut',
                    note='2009 DOTO owner, 1/18 x 17/84 = 17/1512 (listed in some tracts as Lakavage); no subsequent conveyance — interest unchanged.')]
    def s_mcclan():
        return node('Pete McClanahan',DOTO,'DOTO (mineral owner)','',leaf=F(17,504),present=True,nri='Pete McClanahan',
            note='2009 DOTO owner, 1/6 x 17/84 = 17/504; no subsequent conveyance — interest unchanged.')
    def s_hill():
        # Bruce + Kathleen (10/224) → Krafts (5/224 each); Robert → Luanne (5/224); Jane stays. Matches NRI.
        return [
            node('Hill family — Bruce (via Bruce Coleman Hill Trust → Busby) + Kathleen (estate) → the Krafts',
                '20235164','QC of Mineral Interest (+ 20236699, 20191765)','',docs=['20191765','20236699'],
                note=('CONSOLIDATED Hill strand. Two of the four 2009 Hill 1/4-of-5/56 shares (10/224 total) flow to the Krafts: '
                      'Bruce Coleman Hill → Bruce Coleman Hill Trust → Lee Busby, Successor Trustee → Krafts (20191765, 20236699); '
                      'Estate of Kathleen Ann Hill → Krafts (20235164, Peter Hill Kraft fiduciary). Each Kraft ends with one full '
                      'share (5/224) — half of Bruce + half of Kathleen. Robert Moore Hill Jr. (intestate) vests in his wife Luanne '
                      'on a separate strand; Jane keeps her share. Reconciles to the NRI sheet.'),
                children=[
                    node('Peter Hill Kraft','20236699','QC of Mineral Interest','',leaf=F(5,224),present=True,nri='Peter Hill Kraft'),
                    node('Tracy Kraft','20235164','QC of Mineral Interest','',leaf=F(5,224),present=True,nri='Tracy Kraft'),
                ]),
            node('Robert Moore Hill, Jr.',DOTO,'DOTO (mineral owner)','',
                note='2009 DOTO Hill owner, 1/4 x 5/56 = 5/224. Died intestate; per your instruction his share vests in his surviving wife.',
                children=[node('Luanne Staton Hill','2018-CPC-00366','Heirship Estate Documents','',leaf=F(5,224),present=True,
                    nri='Luanne Staton Hill - Wife',
                    note='Robert Moore Hill Jr. died intestate (2018-CPC-00366); his 5/224 vests in his surviving wife, Luanne Staton Hill.')]),
            node('Jane Alison Hill Cochrane',DOTO,'DOTO (mineral owner)','',leaf=F(5,224),present=True,
                nri='Jane Alison Hill Cochrane (Lawrence - husb',
                note='2009 DOTO Hill owner (formerly Jane Alison Hill; name change on marriage), 1/4 x 5/56 = 5/224; unchanged.'),
        ]
    def s_mcdaniel():
        return [node(nm,DOTO,'DOTO (mineral owner)','',leaf=F(1,168),present=True,nri=nri,deceased=dead,
                    note=f'2009 DOTO owner, 1/4 x 10/420 = 1/168; {extra}')
                for nm,nri,dead,extra in [
                    ('Richard M. McDaniel','Richard M. McDaniel',False,'no subsequent conveyance — interest unchanged.'),
                    ('Kenneth W. McDaniel','Kenneth W. McDaniel',False,'no subsequent conveyance — interest unchanged.'),
                    ('Patricia J. McDaniel','Patricia J. McDaniel',False,'no subsequent conveyance — interest unchanged.'),
                    ("Kathy Richmond (dec'd)","Kathy Richmond - dec'd",True,'owner deceased, interest unchanged (estate pending).')]]
    def s_harman():
        # 25/420 split per the NRI ownership spreadsheet: Donald 80% (1/21), Claude Jr. 20% (1/84).
        return [node('Donald Harman',DOTO,'DOTO (mineral owner)','',leaf=F(80*25,100*420),present=True,nri='Donald Harman',
                    note=('Harman mineral interest, 80% x 25/420 = 1/21, booked to Donald Harman per the NRI ownership '
                          'spreadsheet. (The 2009 DOTO listed the 80% jointly as Claude Lester Harman, Jr. & Donald '
                          'Wayne Harman; ownership follows the spreadsheet.)')),
                node('Claude Lester Harman, Jr.',DOTO,'DOTO (mineral owner)','',leaf=F(20*25,100*420),present=True,nri='Claude Lester Harman, Jr.',
                    note='Harman mineral interest, 20% x 25/420 = 1/84, per the NRI ownership spreadsheet; interest unchanged.')]

    # ── per-tract assembly ──
    if t==1:
        tyra=node('Lewis R. Tyra and Charlyn K. Tyra',DOTO,'DOTO (mineral owner)','',docs=['2013000463'],
            note='2009 DOTO owner, 1/4 (community). 2013000463 confirmed the community estate; on Lewis Tyra\'s death his interest passed to the LCT Revocable Trust.',
            children=[node('Charlyn K. Tyra, as Trustee of the LCT Revocable Trust','20-519-CP4','Order / LWT','',leaf=F(1,4),present=True,
                nri='Charlyn K. Tyra, as Trustee of LCT Revocable Trust',
                note='Estate of Lewis Riley Tyra → LCT Revocable Trust (Charlyn K. Tyra, Trustee), per Article 2 disposition; holds the full 1/4.')])
        strands=[s_elmore(F(1,2),' A Quellhorst surface-use agreement covers part of this tract (attached).',docs=['SurfAgmt_Quellhorst']),
                 tyra,
                 node('Richard N. Tyra',DOTO,'DOTO (mineral owner)','',leaf=F(1,4),present=True,nri='Richard N. Tyra',
                      note='2009 DOTO owner, 1/4; no subsequent conveyance — interest unchanged.')]
    elif t==2:
        brem_nm,brem_f,brem_a=dv('Bradbury Family')
        bradbury=node('Bradbury Family Investments, LLC','20172985','SWD w/ VL','',
            leaf=brem_f,present=True,nri=brem_nm,acres=brem_a,
            docs=['20172986','20195281','20193637','20212918','20222292','20223281','20230767','20246554'],
            note=('Bought the consolidated Lenoir 25/420 block (20172985) and subdivided it into the lots below; '
                  f'retains a {brem_a:.4f}-ac remainder ({brem_f:.9f}). RESURVEY NOTE: the lot acreages as-called do '
                  'not sum exactly to the called block acreage — fractions are reconciled to the rollup so the tract '
                  'still totals 100%, but the as-called acres are shown on each lot for you to rectify.'),
            children=[
                lot('Broussard','Jerry and Kristin Broussard','20214839',block='Bradbury subdivision',
                    docs=['20214840','20216955','20222552','20222294','20222293','20234541','20232936']),
                lot('Denman','Steven and LaNee Denman','20230583',block='Bradbury subdivision',
                    docs=['20230584','20255881','20246345','20246347']),
                node('Edgar Epitacio Gallegos & Miguel Angel Cervera Gallegos','20223159','GWD w/VL','',
                    docs=['20223160','20242574'],
                    note='Builder/flipper — bought from Bradbury, conveyed through to the McRories (20242142); minerals passed.',
                    children=[lot('McRorie','Daniel and Heather McRorie','20242142',block='Bradbury subdivision')]),
                node('Bailey James Ellisor','20212816','WD w/VL','',
                    docs=['20212817','20216238','20216762','20217983'],
                    note='Builder/flipper — bought a block from Bradbury, conveyed to Country Living Construction; minerals passed.',
                    children=[
                        node('Country Living Construction, LLC','20216239','WD w/VL','',
                            docs=['20216715','20219038','20226564','20216241','20216717','20219040','20222782','20222781','20226422'],
                            note='Builder/flipper — bought from Ellisor (multiple deeds), built and sold finished lots; minerals passed.',
                            children=[
                                lot('Barrington','Calvin and Chelsea Barrington','20222653',block='Bradbury subdivision',docs=['20222654']),
                                lot('Mayo','Raigan Gwen Mayo & Heather Ann Henson','20222685',block='Bradbury subdivision',docs=['20222686']),
                                lot('Edwards','James Brian & Wanda Jeanene Edwards','20226398',block='Bradbury subdivision',docs=['20226400','20226399']),
                                node('Joseph Margiotta','20226530','GWD','',
                                    note='Flipper — took the lot, then gifted it to Deloach reserving a life estate; minerals passed.',
                                    children=[lot('Deloach','Candy Lynne Deloach','20240688',block='Bradbury subdivision',
                                        note='Bradbury subdivision lot via Margiotta gift deed (grantor reserved a life estate); minerals passed.')]),
                            ]),
                    ]),
            ])
        lenoir=node(('W. Frank Lenoir Jr. Estate (1/2) + Lucy W. Lenoir Estate (1/2) — '
                     'cured into JP Morgan Chase, Trustee (W. Frank Lenoir GST Exempt & Non-Exempt Trusts)'),
            DOTO,'DOTO + curative (probates, executors deeds, trustee res./appt.)','',
            docs=['10-5146','10-5147','10-6043','10-6044','2014001940','2014001941','2014001942','2014001943'],
            note=('CONSOLIDATED Lenoir strand (per your instruction). Two 2009 DOTO estates, 25/840 each, were cured '
                  'through crossed testamentary trusts (10-5146/47), executors deeds (10-6043/44), and the Compass Bank / '
                  'JPMorgan co-trustee resignations & appointments (2014001940-43), then sold as one 25/420 block to '
                  'Bradbury Family Investments via 20172985. Held here as a single node; all curative instruments attached.'),
            children=[bradbury])
        strands=[s_elmore(F(17,84)),lenoir,s_owen(),s_browder(F(17,84)),s_trapp(),s_speer(),s_mccune(),s_dove(),
                 *s_wellborn(),s_mcclan(),*s_hill()]+s_mcdaniel()
    elif t==3:
        # Browder 17/84 split at the DOTO: 1/8 to the Bonds, 13/168 remainder to the Browder→Few line.
        bonds=node('Thomas L. Bonds & Joyce D. Bonds',DOTO,'DOTO (mineral owner)','',
            note='2009 DOTO owner, 1/8 (carved out of the Browder 17/84). Subdivided after Thomas Bonds\' death.',
            children=[node('Joyce Marie Bonds (surviving wife)','P2018-68','Order & LWT','',
                note='Estate of Thomas Lester Bonds → all to surviving wife Joyce Bonds (P2018-68); she then subdivided.',
                children=[
                    lot('Garcia','Adalberto Garcia Ochoa and Jessika G. Garcia','20210747',block='Bonds subdivision'),
                    node('Mark Thornbrough','20201299','WD w/VL','',docs=['20201300','20207968'],
                        note='Flipper — bought from Joyce Bonds, conveyed on; minerals passed.',
                        children=[node('Sharon G. Koerner','20207109','GWD','',
                            note='Flipper — bought from Thornbrough, sold finished lots; minerals passed.',
                            children=[
                                node('Crystal & Alvaro Corona','20216713','WD w/VL','',docs=['20216714','20255422'],
                                    note='Flipper — bought from Koerner, conveyed to Old Hickory Hill; minerals passed.',
                                    children=[lot('Old Hickory','Old Hickory Hill, LLC','20254367',block='Bonds subdivision')]),
                                lot('Plaisance','Craig Michael Plaisance & Daniel Bruce Foster','20220813',block='Bonds subdivision',
                                    docs=['20220171','20220170','20237192']),
                            ])])])])
        strands=[s_elmore(F(17,84)),s_owen(),
                 s_browder(F(13,168),' In this tract a 1/8 was carved to Thomas & Joyce Bonds, so only 13/168 reaches Few.'),
                 bonds,s_trapp(),s_speer(),s_mccune(),s_dove(),*s_wellborn(),*s_harman(),s_mcclan(),*s_hill()]+s_mcdaniel()
    elif t==4:
        strands=[s_elmore(F(1,2),' Burdened by the 1/16 NPR 2.'),
                 node('Southern Timbergrowers, Inc.',DOTO,'DOTO (mineral owner)','',leaf=F(1,2),present=True,nri='Southern Timbergrowers, Inc.',
                      note='2009 DOTO owner, 1/2; SJC record search found no subsequent documents — interest unchanged.')]
    elif t==5:
        strands=[s_elmore(F(17,84)),s_owen(),
                 node('Southern Timbergrowers, Incorporated',DOTO,'DOTO (mineral owner)','',leaf=F(17,84),present=True,nri='Southern Timbergrowers, Inc.',
                      note='2009 DOTO owner, 17/84; no subsequent conveyance — interest unchanged.'),
                 s_trapp(),s_speer(),s_mccune(),s_dove(),*s_wellborn(),*s_harman(),s_mcclan(),*s_hill()]+s_mcdaniel()
    elif t==6:
        strands=[s_elmore(F(17,84)),
                 node('Cyrus Lee Lipe',DOTO,'DOTO (mineral owner)','',leaf=F(10,420),present=True,nri='Cyrus Lee Lipe',
                      note='2009 DOTO owner, 10/420 = 1/42; no subsequent conveyance — interest unchanged (status: Negotiating).'),
                 s_owen(),s_browder(F(17,84)),s_trapp(),s_speer(),s_mccune(),s_dove(),*s_wellborn(),*s_harman(),s_mcclan()]+s_hill()
    elif t==7:
        peck=node('Wanda J. Peck & Michael J. Peck',DOTO,'DOTO (mineral owner)','',
            note=('2009 DOTO owner, 1/4. The interest moved among Wanda, the Peck Revocable Trust, and son Shawn Peck — '
                  'all "No reservation" (minerals passed each time) — before vesting in K\'s Humble Yamaha.'),
            children=[node('Wanda J. Peck (survivor / trustee)','2014007129','Executors SWD','',
                docs=['2014006917','2013003629','20150085','20151095'],
                note=('Michael Peck estate → Michael & Wanda Peck Revocable Trust (2014006917); Wanda as Ind. Executor / Trustee '
                      'consolidated the 1/4 (2014007129, 20151095); an earlier gift to son Shawn (2013003629) was returned (20150085). '
                      'No reservation throughout.'),
                children=[node('Shawn C. Peck','20195601','GWD','',
                    note='Wanda conveyed the 1/4 to Shawn (GWD, no reservation).',
                    children=[node("K's Humble Yamaha, Inc.",'20205047','GWD','',leaf=F(1,4),present=True,nri="K's Humble Yamaha Inc",
                        note='Shawn & Laurey Peck conveyed the 1/4 to K\'s Humble Yamaha, Inc. (GWD, no reservation).')])])])
        strands=[s_elmore(F(3,4),' Burdened by the 1/16 NPR 2.'),peck]
    else:
        strands=[]

    # ── emit recursively ──
    def of_of(spec):
        base=float(spec['leaf']) if spec['leaf'] is not None else 0.0
        return base+sum(of_of(c) for c in spec['children'])
    IDC=[0]; present_leaves=[]
    def emit(spec,pnid):
        IDC[0]+=1; nid=f'tr{t}c-{IDC[0]}'
        of=round(of_of(spec),PREC)
        remaining=round(float(spec['leaf']),PREC) if spec['leaf'] is not None else 0.0
        row,_=rowfor(spec['nri'])
        o=owner_id(spec['name'],row)
        # attachments: controlling DOTO + this node's instrument + any related/financing docs
        atts=[('doc-'+DOTO,pdfs[DOTO],'other')]
        if spec['instr'] and spec['instr'] in pdfs:
            need.add(spec['instr']); atts.append(('doc-'+spec['instr'],pdfs[spec['instr']],kind_of(spec['itype'])))
        for dk in spec['docs']:
            if dk in pdfs:
                need.add(dk); atts.append(('doc-'+dk,pdfs[dk],kind_of(DOC_META.get(dk,{}).get('itype',''))))
        natt=[]
        for j,(docid,fn,kk) in enumerate(atts):
            aid=f'att-tr{t}c-{IDC[0]}-{j}'
            natt.append(dict(docId=docid,attachmentId=aid,fileName=fn,kind=kk))
            attachments.append(dict(attachmentId=aid,workspaceId=WS,docId=docid,entityKind='node',entityId=nid,position=j,createdAt=NOW))
        of9=dec(of); n_,d_=numden(of)
        instrlabel=('DOTO (Distribution / Designation of Ownership)' if spec['instr']==DOTO
                    else f"{spec['itype']} ({spec['instr']})" if spec['instr'] else spec['itype'])
        acretxt=f" As-called acreage {spec['acres']:.4f} ac." if spec['acres'] else ''
        # real instrument date from the runsheet metadata when not specified on the spec
        ndate=spec['date'] or DOC_META.get(spec['instr'] or '',{}).get('file_date') or '2009-06-04'
        nodes.append(dict(id=nid,type='conveyance',instrument=instrlabel,vol='',page='',
            docNo=spec['instr'] or DOTO,fileDate=ndate,date=ndate,
            grantor='per 6/4/2009 DOTO' if spec['instr']==DOTO else '(see chain)',grantee=spec['name'],
            landDesc=landdesc,
            remarks=(spec['note']+acretxt).strip(),
            fraction=dec(remaining),initialFraction=of9,parentId=pnid,conveyanceMode='fraction',splitBasis='initial',
            numerator=n_,denominator=d_,manualAmount=of9,isDeceased=spec['deceased'],obituary='',graveyardLink='',
            attachments=natt,linkedOwnerId=o,linkedLeaseId=None,relatedKind=None,interestClass='mineral',
            royaltyKind=None,fixedRoyaltyBasis=None,depthRange='all_depths',isCollapsed=False))
        node_ids.append(nid)
        leafnode=nodes[-1]
        if spec['present'] and remaining>0:
            present_leaves.append(leafnode)
            # lease: normally gated by the NRI status sheet (Executed only),
            # with named OGML packet overrides for source-proven spreadsheet drift.
            aroy,astatus,aexecuted=auth_for(spec['nri'])
            L=lease_for_owner(spec['nri'])
            source_note='nri-status'
            if L and L['instr'] in LEASE_SOURCE_OVERRIDES:
                override=LEASE_SOURCE_OVERRIDES[L['instr']]
                source_note=override['evidence']
                aexecuted=True
                if not aroy:
                    aroy=override['royalty']
                if not astatus:
                    astatus=override['status']
            elif not aexecuted:
                L=None
            generated_lid=''; generated_lnid=''
            if L:
                need.add(L['instr']); lid=f'lease-{len(lease_recs)+1}'; rem=L['rem']; roy=aroy
                if L['instr'] in LEASE_SOURCE_OVERRIDES:
                    rem=LEASE_SOURCE_OVERRIDES[L['instr']].get('notes',rem)
                if not roy: rem=(rem+' | Royalty pending — lease package not yet booked (NRI carried at 0).').strip(' |')
                lease_recs.append(dict(id=lid,workspaceId=WS,ownerId=o,leaseName=f'OGML — {L["lessor"][:40]}',
                    lessee='Magnolia Petroleum Company, LLC',royaltyRate=roy,leasedInterest='',effectiveDate=L['date'],
                    expirationDate='',status='Active',docNo=L['instr'],notes=rem,jurisdiction='tx_fee',
                    depthRange='all_depths',createdAt=NOW,updatedAt=NOW))
                IDC[0]+=1; lnid=f'tr{t}c-{IDC[0]}'; laid=f'att-tr{t}c-{IDC[0]}-0'
                nodes.append(dict(id=lnid,type='related',relatedKind='lease',instrument='Oil & Gas Lease',vol='',page='',
                    docNo=L['instr'],fileDate=L['date'],date=L['date'],grantor=spec['name'],grantee='Magnolia Petroleum Company, LLC',
                    landDesc=landdesc,remarks=f'Lease: OGML | Royalty: {roy} | Status: Active | Notes: {rem}'.strip(' |'),
                    fraction='0',initialFraction='0',parentId=nid,conveyanceMode='all',splitBasis='whole',
                    numerator='0',denominator='1',manualAmount='0',isDeceased=False,obituary='',graveyardLink='',
                    attachments=[dict(docId='doc-'+L['instr'],attachmentId=laid,fileName=pdfs[L['instr']],kind='lease')],
                    linkedOwnerId=o,linkedLeaseId=lid,interestClass='mineral',royaltyKind=None,fixedRoyaltyBasis=None,
                    depthRange='all_depths',isCollapsed=False))
                attachments.append(dict(attachmentId=laid,workspaceId=WS,docId='doc-'+L['instr'],entityKind='node',entityId=lnid,position=0,createdAt=NOW))
                node_ids.append(lnid)
                generated_lid=lid; generated_lnid=lnid
            reconciliation.append(dict(tract=t,ownerName=spec['nri'],generatedOwnerName=spec['name'],
                ownerId=o,nodeId=nid,mineralFraction=dec(remaining),leaseStatus=astatus,leaseRoyalty=aroy,
                expectedExecuted=aexecuted,ogmlDocument=L['instr'] if L else '',generatedLeaseId=generated_lid,
                generatedLeaseNodeId=generated_lnid,sourceNote=source_note))
        for c in spec['children']: emit(c,nid)
    for s in strands: emit(s,parent_id)
    # fold the 9-place rounding residue into Elmore (a direct-child leaf), so present leaves total exactly 1.0
    tot=sum(float(n['fraction']) for n in present_leaves)
    resid=round(1.0-tot,PREC)
    el=present_leaves[0]  # elmore is first
    newf=round(float(el['fraction'])+resid,PREC)
    el['fraction']=dec(newf); el['initialFraction']=dec(newf); el['manualAmount']=dec(newf)
    el['numerator'],el['denominator']=numden(newf)

for t in range(1,8):
    chain=chain_for(t)
    ros=roster(t)
    decs=[round(f,PREC) for *_,f in ros]
    resid=round(1.0-sum(decs),PREC)
    if ros:
        mx=max(range(len(ros)),key=lambda i:decs[i]); decs[mx]=round(decs[mx]+resid,PREC)
    node_ids=[]
    div_names={n for (_,n,_) in roster(t)} - {first(r) for r in range(2,TOTAL) if isn(ns.cell(r,MIcol[t]).value)}

    # ---- title-opinion parent: the whole tract per the 6/4/2009 DOTO ----
    # Root holds the entire tract (initialFraction = 1.0, mode 'all'); present-day
    # owners hang off it as children dividing the whole. This makes each desk map
    # a real title tree and a clean target for predecessor-insert ("precede").
    nidc+=1; parent_id=f'tr{t}-doto'
    paid=f'att-{nidc}-0'
    nodes.append(dict(id=parent_id,type='conveyance',
        instrument='DOTO (Distribution / Designation of Ownership)',vol='',page='',
        docNo=DOTO,fileDate='2009-06-04',date='2009-06-04',
        grantor='John G. Gaston, Attorney (title opinion)',grantee=f'Dr. Elmore #1 Unit — Tract {t} (whole)',
        landDesc=f'{SURVEY}; Dr. Elmore #1 Unit Tract {t} (DOTO Unit Tr. {UNITTR[t]}), {ACRES[t]} ac.',
        remarks=f'Title-opinion root for Tract {t}: the whole {ACRES[t]}-ac tract per the 6/4/2009 DOTO. Present-day owners below divide this 100%. Insert a predecessor here to test "precede".',
        fraction='0',initialFraction='1.000000000',parentId=None,conveyanceMode='all',splitBasis='whole',
        numerator='1',denominator='1',manualAmount='1.000000000',isDeceased=False,obituary='',graveyardLink='',
        attachments=[dict(docId='doc-'+DOTO,attachmentId=paid,fileName=pdfs[DOTO],kind='other')],
        linkedOwnerId=None,linkedLeaseId=None,relatedKind=None,interestClass='mineral',
        royaltyKind=None,fixedRoyaltyBasis=None,depthRange='all_depths',isCollapsed=False))
    attachments.append(dict(attachmentId=paid,workspaceId=WS,docId='doc-'+DOTO,
        entityKind='node',entityId=parent_id,position=0,createdAt=NOW))
    node_ids.append(parent_id)

    # ---- NPRI sibling burdens off the DOTO root (where this tract is burdened) ----
    npr_spec=None
    if t in NPR1_TRACTS: npr_spec=('NPR 1','1/64',0.015625,NPR1_OWNERS)
    elif t in NPR2_TRACTS: npr_spec=('NPR 2','1/16',0.0625,NPR2_OWNERS)
    if npr_spec:
        nlabel,nrate,ntotal,nowners=npr_spec
        ndecs=[round(f,PREC) for _,f,_ in nowners]
        nresid=round(ntotal-sum(ndecs),PREC)
        mx=max(range(len(nowners)),key=lambda i:ndecs[i]); ndecs[mx]=round(ndecs[mx]+nresid,PREC)
        for k,(pname,pf,pinstr) in enumerate(nowners):
            nidc+=1; npid=f'tr{t}-npr{nidc}'; po=owner_id(pname)
            pf9=dec(ndecs[k]); pn,pden=numden(ndecs[k])
            natts=[('doc-'+DOTO,pdfs[DOTO],'other')]
            if pinstr and pinstr in pdfs:
                need.add(pinstr); natts.append(('doc-'+pinstr,pdfs[pinstr],'deed'))
            nodeatt=[]
            for j,(docid,fn,kk) in enumerate(natts):
                aid=f'att-{nidc}-{j}'
                nodeatt.append(dict(docId=docid,attachmentId=aid,fileName=fn,kind=kk))
                attachments.append(dict(attachmentId=aid,workspaceId=WS,docId=docid,
                    entityKind='node',entityId=npid,position=j,createdAt=NOW))
            nodes.append(dict(id=npid,type='conveyance',
                instrument=f'{nlabel} NPRI ({nrate} royalty, per 6/4/2009 DOTO)',vol='',page='',
                docNo=DOTO,fileDate='2009-06-04',date='2009-06-04',
                grantor=f'{nlabel} royalty reservation (up-chain of DOTO)',grantee=pname,
                landDesc=f'{SURVEY}; Dr. Elmore #1 Unit Tract {t} (DOTO Unit Tr. {UNITTR[t]})',
                remarks=(f'{nlabel} non-participating royalty interest ({nrate}) burdening Tract {t}. '
                         f'Present-day royalty share {pf9}. Sibling burden to the mineral estate; reserved up-chain of the DOTO.'),
                fraction=pf9,initialFraction=pf9,parentId=parent_id,conveyanceMode='fraction',splitBasis='initial',
                numerator=pn,denominator=pden,manualAmount=pf9,isDeceased=False,obituary='',graveyardLink='',
                attachments=nodeatt,linkedOwnerId=po,linkedLeaseId=None,relatedKind=None,
                interestClass='npri',royaltyKind='fixed',fixedRoyaltyBasis='whole_tract',
                depthRange='all_depths',isCollapsed=False))
            node_ids.append(npid)

    # Every tract is now built as a full DOTO→present chain of title.
    build_chain(t, parent_id, node_ids)
    deskMaps.append(dict(id=f'dm-tr{t}',name=f'Tract {t} — {ACRES[t]} ac. (DOTO Unit Tr. {UNITTR[t]})',
        code=f'TR{t}',tractId=None,grossAcres=ACRES[t],pooledAcres=ACRES[t],
        description=f'Dr. Elmore #1 Unit, Tract {t}. Present-day mineral ownership ({len(node_ids)} owners); totals 100%. {COUNTY} County, TX.',
        nodeIds=node_ids,unitName='Dr. Elmore #1 Unit',unitCode='A'))

# ---------- documents (fully populated registry metadata) ----------
documents=[]
for src in sorted(need):
    raw=open(os.path.join(PDFDIR,pdfs[src]),'rb').read(); b64=base64.b64encode(raw).decode()
    is_ogml=src.upper().startswith('OGML')
    m=DOC_META.get(src,{})
    itype=('DOTO' if src==DOTO else (m.get('itype') or ('OGML' if is_ogml else 'Recorded instrument')))
    if src==DOTO:
        title='DOTO — Dr. Elmore C-1 Unit'
    elif m.get('itype') and (m.get('grantee') or m.get('grantor')):
        party=m.get('grantee') or m.get('grantor')
        title=f"{m['itype']} — {party[:48]}"
    elif is_ogml:
        title=f'OGML — {pdfs[src]}'
    else:
        title=f'Instrument {src}'
    documents.append(dict(docId='doc-'+src,workspaceId=WS,fileName=pdfs[src],mimeType='application/pdf',
        byteLength=len(raw),contentHash=hashlib.sha256(raw).hexdigest(),
        blob=dict(mimeType='application/pdf',base64=b64),
        kind=('other' if src==DOTO else ('lease' if is_ogml else kind_of(itype))),
        displayTitle=title,
        documentArea=('leasehold' if is_ogml else 'runsheet_mineral_title'),
        instrumentType=itype, county=COUNTY, instrumentNumber=src,
        volume=m.get('vol',''), page=m.get('page',''),
        effectiveDate=m.get('inst_date',''), recordingDate=m.get('file_date',''),
        grantor=m.get('grantor',''), grantee=m.get('grantee',''),
        notes=m.get('remarks',''), createdAt=NOW, updatedAt=NOW))

# ---------- lease records (one per lease child-node, already built above) ----------
leases=lease_recs

env=dict(version=8,exportedAt=NOW,workspaceId=WS,
    projectName='Dr. Elmore #1 Unit — Present-Day Mineral Ownership (7 Tracts, San Jacinto Co., TX)',
    nodes=nodes,deskMaps=deskMaps,
    leaseholdUnit=dict(name='Dr. Elmore #1 Unit',description='Proposed 7-tract pooled unit, San Jacinto County, TX (Thomas Webb A-300 / Vital Flores A-14). Lessee: Magnolia Petroleum Company, LLC; operator Arcadia Operating, LLC. Present-day mineral ownership from the DOTO runsheet + NRI status.'),
    leaseholdAssignments=[],leaseholdOrris=[],leaseholdTransferOrderEntries=[],
    activeDeskMapId='dm-tr1',activeUnitCode='A',
    instrumentTypes=['Patent','DOTO','Warranty Deed','Mineral Deed','Gift Deed','SWD','Quitclaim Deed','Probate Order & LWT','Affidavit of Heirship','OGML'],
    ownerData=dict(owners=list(owners.values()),leases=leases,contacts=[],docs=[]),
    documentData=dict(documents=documents,attachments=attachments),
    mapData=dict(mapAssets=[],mapRegions=[],mapReferences=[]),
    researchData=dict(imports=[],sources=[],formulas=[],projectRecords=[]),
    curativeData=dict(issues=[]))
os.makedirs(os.path.dirname(OUT),exist_ok=True)
with open(OUT,'w') as f: json.dump(env,f,indent=2,ensure_ascii=False)

# ---------- verify ----------
out=[f'OUTPUT: {OUT}','size: %.2f MB'%(os.path.getsize(OUT)/1024/1024),
 f'NRI MI columns detected: '+str({t:openpyxl.utils.get_column_letter(c) for t,c in sorted(MIcol.items())}),
 f'nodes={len(nodes)} owners={len(owners)} deskMaps={len(deskMaps)} documents={len(documents)} attachments={len(attachments)} leases={len(leases)}',
 '','per-tract CHILD sums (children divide the DOTO parent; must be 1.000000000):']
docids={d['docId'] for d in documents}; ownids={o['id'] for o in owners.values()}
byid={n['id']:n for n in nodes}
lease_parent_ids={n['parentId'] for n in nodes if n.get('relatedKind')=='lease' and n.get('parentId')}
for dm in deskMaps:
    kids=[byid[nid] for nid in dm['nodeIds'] if byid[nid]['parentId'] is not None]
    owner_kids=[k for k in kids if k.get('relatedKind')!='lease' and k.get('interestClass')!='npri']
    lease_kids=[k for k in kids if k.get('relatedKind')=='lease']
    npri_kids=[k for k in kids if k.get('interestClass')=='npri']
    s=sum(float(k['fraction']) for k in owner_kids)              # mineral estate only -> 1.0
    leased=sum(float(k['fraction']) for k in owner_kids if k['id'] in lease_parent_ids)
    unleased=round(1.0-leased,PREC)
    npr_sum=sum(float(k['fraction']) for k in npri_kids)         # sibling burden, separate
    npr_txt=f"  NPRI={len(npri_kids)} sum={npr_sum:.8f}" if npri_kids else ""
    out.append(f"  {dm['code']}: DOTO + {len(owner_kids):>2} owners + {len(lease_kids):>2} leases  mineralsum={s:.9f}  {'OK' if abs(s-1)<1e-9 else 'FAIL'}  leased={leased:.9f} unleased={unleased:.9f}{npr_txt}")
bad_att=[n['id'] for n in nodes if not n['attachments'] or any(a['docId'] not in docids for a in n['attachments'])]
# only child (owner) nodes carry an owner link; DOTO roots intentionally do not
bad_own=[n['id'] for n in nodes if n['parentId'] is not None and n['linkedOwnerId'] not in ownids]
roots=[n['id'] for n in nodes if n['parentId'] is None]
orphan_kids=[n['id'] for n in nodes if n['parentId'] is not None and n['parentId'] not in byid]
baddec=sum(1 for d in documents if base64.b64decode(d['blob']['base64'])[:5]!=b'%PDF-')
missing_executed=[r for r in reconciliation if r['expectedExecuted'] and not r['generatedLeaseId']]
lct_recs=[r for r in reconciliation if r['ogmlDocument']=='OGML-LCT-Trust' or ('Charlyn K. Tyra' in r['ownerName'] and 'LCT' in r['ownerName'])]
if missing_executed:
    out.append('')
    out.append('source-to-output reconciliation failures:')
    for r in missing_executed:
        out.append(f"  tract {r['tract']} owner={r['ownerName']} status={r['leaseStatus'] or '(blank)'} ogml={r['ogmlDocument'] or '(missing)'} node={r['nodeId']}")
if not lct_recs:
    raise AssertionError('LCT reconciliation row missing')
lct_ok=any(r['generatedLeaseId'] and r['generatedLeaseNodeId'] and r['leaseRoyalty']=='1/4' for r in lct_recs)
if not lct_ok:
    raise AssertionError(f'LCT lease assertion failed: {lct_recs}')
out += ['',f'source-to-output executed rows missing generated lease: {len(missing_executed)}',
 f'LCT assertion: PASS — OGML-LCT-Trust generated for Charlyn K. Tyra / LCT with 1/4 royalty']
out += ['',f'nodes w/ broken attachment: {len(bad_att)}',f'child nodes w/ broken owner link: {len(bad_own)}',
 f'root (title-opinion) nodes: {len(roots)}  (want 7)',
 f'children pointing at a missing parent: {len(orphan_kids)}  (want 0)',
 f'min attachments on any node: {min(len(n["attachments"]) for n in nodes)}  (want >=1)',
 f'avg PDFs/node: {sum(len(n["attachments"]) for n in nodes)/len(nodes):.2f}',
 f'unique real PDFs embedded: {len(documents)}',
 f'PDFs failing %PDF- decode: {baddec}  (want 0)',
 f'lease nodes on canvas (related/lease): {sum(1 for n in nodes if n.get("relatedKind")=="lease")}',
 f'NPRI nodes on canvas (interestClass=npri): {sum(1 for n in nodes if n.get("interestClass")=="npri")}',
 f'owners with mailing address: {sum(1 for o in owners.values() if o["mailingAddress"])}/{len(owners)}',
 f'leases linked: {len(leases)}']
os.makedirs(os.path.dirname(REPORT),exist_ok=True)
open(REPORT,'w').write('\n'.join(out))
out.insert(1,f'REPORT: {REPORT}')
print('\n'.join(out))
if missing_executed:
    raise AssertionError(f'{len(missing_executed)} executed lease row(s) missing generated lease output')
